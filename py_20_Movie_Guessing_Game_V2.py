''' Movie Guessing Game V2 - Excel + Selenium
- pick a random title + it`s release year from the MoviePY excel sheet
- look for it in a search engine combined with a movie database title
- open the first match = open the title`s movie database site
- take the synopsys/plot of the movie and display it for the user
- ask the user to guess the title
- give a hint/help(director, release year, stars..) if the user needed '''

import random
from openpyxl import Workbook, load_workbook
wb = load_workbook('MoviePY.xlsx', data_only=True)
ws = wb.active

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


#GAME - BANNER
print()
k = 50
print('*'*k)    
print(' Please wait, the program is loading '.center(k,'*'))
print('*'*k)

round = 0
score = 0

while True:
# VALUE EXTRACTION FROM THE EXCEL
    cellnumber = random.randrange(6,6951)

    cell = 'C' + str(cellnumber)
    movietitle = ws[cell].value

    while movietitle == None or movietitle == '-':        # If it picked a empty cell(1 title in at least 3 merged cells), 
        cellnumber += 1              # looking for the next valid non-empty one
        cell = 'C' + str(cellnumber)
        movietitle = ws[cell].value

    cellRYear = 'E' + str(cellnumber)
    releaseYear = ws[cellRYear].value

    cellHSeen = 'N' + str(cellnumber)
    haveSeen = ws[cellHSeen].value

# VALUE EXTRACTION FROM WEBSITE
    PATH = 'C:\Program Files (x86)\chromedriver.exe'
    driver = webdriver.Chrome(PATH)
    driver.get('https://duckduckgo.com/')  # using search engine as a pre step
                                           # because apart from IMDb, not able to search by movie title + release year
    try:
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, 'q' ))
        )

    except:
        driver.quit()
    
    movieDataBase = 'rotten tomatoes'
    searchForMovie = ' '.join([ movieDataBase, str(movietitle), str(releaseYear) ])

# LOOKING FOR THE MOVIE VIA SEARCH ENGINE
    search = driver.find_element(By.NAME,'q')
    search.send_keys(searchForMovie)
    search.send_keys(Keys.RETURN)

# CLICKING ON THE FIRST RESULT
    search = driver.find_element(By.ID,'r1-0')
    search.click()

# COLLECTING THE INFORMATION FOR THE MOVIE FROM THE MOVIE DATABASE
    plot = driver.find_element(By.ID, 'movieSynopsis').text
    director = driver.find_element(By.CSS_SELECTOR, 'li.meta-row:nth-child(4) > div:nth-child(2) > a:nth-child(1)').text
    star_1 = driver.find_element(By.CSS_SELECTOR, 'div.cast-item:nth-child(1) > div:nth-child(2) > a:nth-child(1) > span:nth-child(1)').text
    star_2 = driver.find_element(By.CSS_SELECTOR, 'div.cast-item:nth-child(2) > div:nth-child(2) > a:nth-child(1) > span:nth-child(1)').text
    star_3 = driver.find_element(By.CSS_SELECTOR, 'div.cast-item:nth-child(3) > div:nth-child(2) > a:nth-child(1) > span:nth-child(1)').text
    driver.quit()

# - THE GAME -
#ROUND - BANNER
    round +=1
    print('\n')
    print((' ROUND ' + str(round) + ' ').center(50, '-'))
    print()
    print('Guess the movie where..')
    print(plot)
    print()

#USER ANSWER
    print('Type:\n"M" for more information\n"R" to reveal the movie title\n"Q" for leave the game')
    answer = input('Your take: ').lower()
    
    if answer == 'q':
        print()
        print('Thx for playing!')
        print()
        break

    if answer == 'r':
        print()
        print('The movie was: ' + str(movietitle).strip())

    if answer == 'm':
        print()
        print('A little help:')
        print('Year of release: ' + str(releaseYear))
        print('Director: ' + director)
        print('Star(s): ' + star_1 + ', ' + star_2 + ', ' + star_3)
        if haveSeen == 1: 
            print('You have seen this movie only once since 05/2012.')
        else:
            print('You have seen this movie ' + str(haveSeen) +  ' times since 05/2012.')
        print()
        answer = input('Your take: ').lower()

    print()
    bePart = 0  # variable for a partially correct answers

    if answer == str(movietitle).lower():
        print('That is correct!')
        score += 1
    else:
        for i in answer.split():                # creating a list from the answer, checking the items in the movie title
            if i in str(movietitle).lower().split():
                bePart += 1
        if bePart >= len((str(movietitle).split())) * 0.5:      # answer matching by 50% of he movie title  - scenario
            print('Almost! The movie was: ' + str(movietitle))
            score += 0.5
        elif answer != 'r':
            print('Incorrect! The movie was: ' + str(movietitle))

    print('Your score: ' + str(score) + '/' + str(round))
    print()
    print('Type:\n"Q" for leave the game\n"N" for the next round')
    answer = input('Next step: ').lower()
    while answer not in ['q', 'n']:
        print()
        print('Please try again.')
        answer = input('Next step: ').lower()
    
    if answer == 'q':
        print()
        print('Thx for playing!')
        print()
        break
    else:
        continue